"""
Genişletilmiş IE-Case seti testleri.

Kapsam:
    * 6 problem seti tam yüklenmeli
    * Fixture dosyaları (xlsx, csv, sqlite) yaratılmış ve okunabilir
    * Her problem için ground truth (analitik veya elle çözülmüş) feasibility
      fonksiyonuyla tutarlı olmalı — regresyon önleyici
    * DataBundle uyumu: iesolver.io.data_loader.load_data her dosyayı sorunsuz
      okumalı (public API üzerinden)
    * NO_CODE problemi objective_value=None ile doğru işaretlenmeli
"""

from __future__ import annotations

import math
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from ie_eval.datasets.ie_case import ie_case_dataset
from ie_eval.validator import check_feasibility


# =============================================================================
# Set boyutu ve metadata
# =============================================================================
def test_dataset_contains_six_problems():
    problems = list(ie_case_dataset.load())
    assert len(problems) == 6
    ids = {p.id for p in problems}
    assert ids == {
        "eoq-basic",
        "transport-2x3",
        "multi-product-inventory",
        "transport-3x2-csv",
        "assignment-3x3-sqlite",
        "abc-classification",
    }


def test_all_problems_carry_ie_case_metadata():
    for p in ie_case_dataset.load():
        assert p.metadata["benchmark"] == "IE-Case"
        assert "problem_type" in p.metadata
        assert "expected_path" in p.metadata
        assert "data_format" in p.metadata


def test_expected_path_distribution():
    """5 CODE + 1 NO_CODE (ABC classification)."""
    paths = [p.metadata["expected_path"] for p in ie_case_dataset.load()]
    assert paths.count("CODE") == 5
    assert paths.count("NO_CODE") == 1


def test_data_format_coverage():
    """DataBundle argümanı için 4 farklı format kanıtı."""
    fmts = {p.metadata["data_format"] for p in ie_case_dataset.load()}
    assert fmts == {"none", "xlsx", "csv", "sqlite"}


# =============================================================================
# Fixture dosyaları var mı, okunabiliyor mu?
# =============================================================================
@pytest.fixture
def problems_by_id():
    return {p.id: p for p in ie_case_dataset.load()}


def test_multi_product_xlsx_exists_and_readable(problems_by_id):
    p = problems_by_id["multi-product-inventory"]
    assert p.data_path is not None
    assert p.data_path.exists()

    wb = openpyxl.load_workbook(p.data_path, read_only=True)
    assert "Products" in wb.sheetnames
    assert "UnitPrices" in wb.sheetnames

    ws = wb["Products"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["product", "demand", "ordering_cost", "holding_cost"]

    products = {r[0].value for r in ws.iter_rows(min_row=2)}
    assert products == {"A", "B", "C"}
    wb.close()


def test_transportation_csv_exists_and_readable(problems_by_id):
    p = problems_by_id["transport-3x2-csv"]
    assert p.data_path is not None
    assert p.data_path.exists()

    lines = p.data_path.read_text(encoding="utf-8").strip().splitlines()
    assert lines[0] == "source,destination,cost,supply,demand"
    assert len(lines) == 1 + 6   # header + 3×2 pairs


def test_assignment_sqlite_exists_and_queryable(problems_by_id):
    p = problems_by_id["assignment-3x3-sqlite"]
    assert p.data_path is not None
    assert p.data_path.exists()

    conn = sqlite3.connect(str(p.data_path))
    rows = conn.execute("SELECT worker, task, cost FROM assignment_costs").fetchall()
    conn.close()
    assert len(rows) == 9  # 3×3


# =============================================================================
# iesolver DataBundle uyumu — public API load_data'yı çağırıyoruz
# =============================================================================
def test_databundle_reads_multi_product_xlsx(problems_by_id):
    """iesolver.io.data_loader public bir API değil ama iç kullanım; buradan
    reach etmek yerine intake_node akışını dolaylı test edeceğiz.
    Bu spot test yalnızca dosyanın loader-uyumlu olduğunu doğrular."""
    from iesolver.io.data_loader import load_data
    p = problems_by_id["multi-product-inventory"]
    bundle = load_data(p.data_path)
    assert bundle.source_type == "xlsx"
    assert "Products" in bundle.tables
    # Products DataFrame'i 3 ürün içermeli
    df = bundle.tables["Products"]
    assert len(df) == 3
    assert set(df["product"]) == {"A", "B", "C"}


def test_databundle_reads_transportation_csv(problems_by_id):
    from iesolver.io.data_loader import load_data
    p = problems_by_id["transport-3x2-csv"]
    bundle = load_data(p.data_path)
    assert bundle.source_type == "csv"
    assert "data" in bundle.tables
    df = bundle.tables["data"]
    assert len(df) == 6
    assert set(df["source"]) == {"S1", "S2", "S3"}


def test_databundle_reads_assignment_sqlite(problems_by_id):
    from iesolver.io.data_loader import load_data
    p = problems_by_id["assignment-3x3-sqlite"]
    bundle = load_data(p.data_path)
    assert bundle.source_type == "sqlite"
    assert "assignment_costs" in bundle.tables
    df = bundle.tables["assignment_costs"]
    assert len(df) == 9
    assert set(df["worker"]) == {"W1", "W2", "W3"}


# =============================================================================
# Ground truth doğrulaması — solution feasibility fonksiyonuna uymalı
# =============================================================================
def _assert_ground_truth_feasible(problem):
    """Kayıtlı ground truth solution kendi feasibility_fn'ini geçmeli.

    Bu bir tutarlılık invariant'ı — problem tanımı kendinden çelişkili değil.
    """
    if problem.ground_truth.feasibility_fn is None:
        return
    result = check_feasibility(problem.ground_truth.solution, problem.ground_truth)
    assert result.feasible, (
        f"{problem.id}: kayıtlı ground truth solution kendi feasibility check'ini "
        f"geçemedi — violations: {result.violations}"
    )


def test_all_ground_truth_solutions_are_feasible():
    for p in ie_case_dataset.load():
        _assert_ground_truth_feasible(p)


# =============================================================================
# Problem-özgü ground truth sayısal doğrulama
# =============================================================================
def test_multi_product_expected_tac(problems_by_id):
    """Analitik: TAC = sum(sqrt(2*D_i*S_i*H_i)) — hesaplayıp karşılaştır."""
    p = problems_by_id["multi-product-inventory"]
    expected = (
        math.sqrt(2 * 1200 * 50 * 1.0)
        + math.sqrt(2 * 800 * 40 * 2.0)
        + math.sqrt(2 * 1500 * 25 * 0.5)
    )
    assert p.ground_truth.objective_value == pytest.approx(expected, rel=1e-9)


def test_multi_product_solution_matches_formula(problems_by_id):
    """Her ürün Q_i* = sqrt(2*D_i*S_i/H_i)."""
    p = problems_by_id["multi-product-inventory"]
    sol = p.ground_truth.solution
    assert sol["Q_A"] == pytest.approx(math.sqrt(2 * 1200 * 50 / 1.0))
    assert sol["Q_B"] == pytest.approx(math.sqrt(2 * 800 * 40 / 2.0))
    assert sol["Q_C"] == pytest.approx(math.sqrt(2 * 1500 * 25 / 0.5))


def test_transport_csv_optimal_is_500(problems_by_id):
    """MODI ile doğrulanmış: 500."""
    p = problems_by_id["transport-3x2-csv"]
    assert p.ground_truth.objective_value == 500.0
    sol = p.ground_truth.solution
    # Verilen çözümün toplam maliyeti = 500 (independent recompute)
    cost_matrix = {
        ("S1", "D1"): 4, ("S1", "D2"): 6,
        ("S2", "D1"): 5, ("S2", "D2"): 3,
        ("S3", "D1"): 7, ("S3", "D2"): 2,
    }
    total = 0.0
    for (src, dst), c in cost_matrix.items():
        total += c * sol[f"x_{src}_{dst}"]
    assert total == 500.0


def test_assignment_optimal_is_9(problems_by_id):
    """Hungarian/brute-force ile doğrulanmış: 9."""
    p = problems_by_id["assignment-3x3-sqlite"]
    assert p.ground_truth.objective_value == 9.0

    # Bağımsız cost matrisi ile hesapla
    cost = {
        ("W1", "T1"): 9, ("W1", "T2"): 2, ("W1", "T3"): 7,
        ("W2", "T1"): 6, ("W2", "T2"): 4, ("W2", "T3"): 3,
        ("W3", "T1"): 5, ("W3", "T2"): 8, ("W3", "T3"): 1,
    }
    total = 0.0
    for (w, t), c in cost.items():
        total += c * p.ground_truth.solution[f"x_{w}_{t}"]
    assert total == 9.0


# =============================================================================
# NO_CODE problem — sayısal metrikler skip
# =============================================================================
def test_abc_classification_is_no_code(problems_by_id):
    p = problems_by_id["abc-classification"]
    assert p.metadata["expected_path"] == "NO_CODE"
    assert p.ground_truth.objective_value is None
    assert p.ground_truth.feasibility_fn is None
    assert p.metadata.get("expert_review_required") is True


# =============================================================================
# Fixture idempotency — modül tekrar yüklemek dosyayı bozmamalı
# =============================================================================
def test_fixtures_are_idempotent(tmp_path, monkeypatch):
    """_ensure_fixtures ikinci çağrıda mevcut dosyayı ezmemeli."""
    from ie_eval.datasets import ie_case as ie_case_mod

    # Aynı dizinde ikinci kez çağır → hata olmamalı
    paths_1 = ie_case_mod._ensure_fixtures()
    paths_2 = ie_case_mod._ensure_fixtures()
    for key in paths_1:
        assert paths_1[key] == paths_2[key]
        assert paths_1[key].exists()


# =============================================================================
# metadata_filter ile IE-Case altkümesi — analiz akışıyla uyum
# =============================================================================
def test_metadata_filter_isolates_data_format():
    """DataBundle argümanı kanıtı: xlsx/csv/sqlite formatlarını ayrı analizle."""
    problems = list(ie_case_dataset.load())
    xlsx_only = [p for p in problems if p.metadata["data_format"] == "xlsx"]
    csv_only = [p for p in problems if p.metadata["data_format"] == "csv"]
    sqlite_only = [p for p in problems if p.metadata["data_format"] == "sqlite"]
    assert len(xlsx_only) == 1
    assert len(csv_only) == 1
    assert len(sqlite_only) == 1


def test_metadata_filter_isolates_no_code():
    problems = list(ie_case_dataset.load())
    no_code = [p for p in problems if p.metadata["expected_path"] == "NO_CODE"]
    assert len(no_code) == 1
    assert no_code[0].id == "abc-classification"
